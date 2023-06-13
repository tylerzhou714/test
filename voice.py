from openpyxl import load_workbook
from aip import AipSpeech
import playaudio

def playScore(inputName):
    # 百度AIP的配置信息
    APP_ID = '32891123'
    API_KEY = 'rYQN28Dmv8zRusPxn4LQfuO8'
    SECRET_KEY = 'admRPxzMPYxRuHWOvt81zeDQCjtqUFNq'

    # 初始化百度AIP客户端
    client = AipSpeech(APP_ID, API_KEY, SECRET_KEY)

    # 加载Excel文件
    wb = load_workbook('data/cjcx.xlsx')
    sheet = wb.active

    # 读取Excel中的数据并进行语音合成
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        score = row[2]
        if name == inputName:
            result = client.synthesis(f'{name}的成绩是{score}分', 'zh', 1, {
                'vol': 5,
                'spd': 5,
                'pit': 5,
                'per': 4
            })

            if not isinstance(result, dict):
                # 保存语音合成结果为mp3文件
                filename = f'data/{name}_score.mp3'
                with open(filename, 'wb') as f:
                    f.write(result)
                print(f'{name}的成绩已保存为{filename}')

                # 使用playaudio播放成绩
                playaudio.playaudio(filename)
                # 关闭Excel文件
                wb.close()
                return score

    print(f'未找到学生{name}的成绩。')
    # 关闭Excel文件
    wb.close()
    return None
